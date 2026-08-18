import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']  # this will return the first sheet incase there are many
    # cell1 = sheet['a1']  # this is used to return the first cell
    # cell2 = sheet.cell(1, 1) # it will return the value of the given cell coordinates
    # print(cell1.value)  # this both works with the first cell1 and cell2

    # print(f"Total rows: {sheet.max_row}") # displays the total number of rows on the excell sheet
    # print(f"Total columns: {sheet.max_column}") # displays the total number of columns on the excell sheet

    for row in range(2, sheet.max_row + 1):  # the 2 is the row you wna to display for 1 is the first row on the sheet
        cell = sheet.cell(row, 3)  # the column on the sheet
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
